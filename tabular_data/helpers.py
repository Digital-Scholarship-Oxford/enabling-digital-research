import os
import re
import math
import pandas as pd
import xml.etree.ElementTree as ET
from pathlib import Path
from tqdm import tqdm
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.comments import Comment
import elementpath

# Function to import xml files and their config files
def import_files(xml_path, config_path, xml_recursive=False, config_recursive=False):
    """
    Imports and processes authority files and their configuration.
    Args:
        xml_path (str): Directory containing XML files.
        config_path (str): Directory containing authority configuration CSV files.
    Returns:
        tuple: A dictionary of parsed XML files, a dictionary of configuration DataFrames, and a dictionary of empty DataFrames.
    """
    # Read and parse XML files
    xml_files = read_files(xml_path, pattern=".xml", recursive=xml_recursive)

    print(xml_files)
    xml_data = {}
    for file in tqdm(xml_files, desc="Parsing XML files"):
        try:
            filename = os.path.splitext(os.path.basename(file))[0]
            xml_data[filename] = parse_xml(file)
        except Exception as e:
            tqdm.write(f"Failed to parse file {file}. Error: {e}")

    # Read and parse CSV configuration files
    config_files = sorted(read_files(config_path, pattern=".csv", recursive=config_recursive))
    config_list = {}
    for file in tqdm(config_files, desc="Parsing config files"):
        try:
            name = os.path.splitext(os.path.basename(file))[0]
            config_list[name] = pd.read_csv(file, dtype='str', na_values=["", "nan", "NaN"]).where(pd.notna, None)
        except Exception as e:
            tqdm.write(f"Failed to parse config file {file}. Error: {e}")

    # Create an empty DataFrame for each configuration file
    df_list = {
        name: pd.DataFrame(columns=config['section'] + ": " + config['heading'])
        for name, config in config_list.items()
    }

    return xml_data, config_list, df_list

# Helper function to read files from a directory
def read_files(directory, pattern, recursive=True):
    """
    Reads files from a specified directory.
    Args:
        directory (str): Directory to search for files.
        pattern (str): File extension pattern to match.
        recursive (bool): Whether to search subdirectories.
    Returns:
        list: List of file paths.
    """
    WORKING_DIR = Path(__file__).resolve().parents[1] / 'tabular_data'

    try:
        directory_path = WORKING_DIR / directory
        if recursive:
            files = list(directory_path.rglob(f"*{pattern}"))
        else:
            files = list(directory_path.glob(f"*{pattern}"))
        return [str(file) for file in files]
    except Exception as e:
        tqdm.write(f"Reading files in {directory} failed. Error: {e}")
        raise

# Helper function to parse an XML file and return its root element
def parse_xml(file):
    """
    Parses an XML file and returns the root element.
    Args:
        file (str): Path to the XML file.
    Returns:
        Element: The root element of the parsed XML file.
    """
    try:
        tree = ET.parse(file)
        root = tree.getroot()
        return root
    except Exception as e:
        tqdm.write(f"Parsing {file} failed. Error: {e}")
        raise

# Function to extract data from the XML files based on the configuration files
def process_file(
    file_type,
    config_name,
    config,
    xml_data,
    df_list,
    csv_output_dir,
    json_output_dir,
    separator_map=None,
    lookup_df_list=None,
    bar_pos=1,
    cores_spare=0
):
    """
    Function that processes either authority or collection files depending on file_type.

    Args:
        file_type (str): Either 'authority' or 'collection' to choose which branch to run.
        config_name (str): Name of the configuration file.
        config (DataFrame): The configuration DataFrame.
        xml_data (dict): Dictionary of XML data.
        df_list (dict): Dictionary of DataFrames keyed by configuration name.
        csv_output_dir (str): Output directory for CSV files.
        json_output_dir (str): Output directory for JSON files.
        separator_map (dict): Dictionary of separators for authority lookups. Default None.
        lookup_df_list (dict): Dictionary of DataFrames for authority XML files; used for collection branch. Default None.
        bar_pos (int): Position parameter for tqdm progress bar. Default 1.
    Returns:
        tuple: (config_name, processed DataFrame)
    """
    # Load the DataFrame
    df = df_list[config_name]

    if file_type == "authority":
        # Extract columns for authority processing
        try:
            auth_files, xpaths = (
                config[col].tolist() for col in ["auth_file", "xpath"]
            )
        except Exception as e:
            tqdm.write(f"Failed to extract configuration columns for '{config_name}'. Error: {e}")
            return config_name, df

        # Count cores
        num_workers = os.cpu_count() - cores_spare or 1

        # Prepare arguments
        all_args = [
            (i, xpath, auth_file)
            for i, (xpath, auth_file)
            in enumerate(zip(xpaths, auth_files))
        ]

        # Split into batches
        max_batches = num_workers + 1
        batch_size = max(1, math.ceil(len(all_args) / max_batches))

        batches = [
            all_args[i : i + batch_size]
            for i in range(0, len(all_args), batch_size)
        ]

        # Dispatch batches in parallel
        with ProcessPoolExecutor(max_workers=num_workers) as executor:
            futures = [
                executor.submit(process_authority_batch, batch, xml_data)
                for batch in batches
            ]
            for future in tqdm(
                as_completed(futures),
                total=len(futures),
                desc=f"File '{config_name}'",
                position=bar_pos
            ):
                batch_result = future.result()
                for i, col_data in batch_result.items():
                    df.iloc[:, i] = col_data

        # Defragment the DataFrame by concatenation
        df = pd.concat([df], ignore_index=True)

        # Sort authority data
        df = sort_authority_df(df)

    elif file_type == "collection":
        # Extract columns for collection processing
        try:
            xpaths, auth_files, auth_sections, auth_cols, separators = (
                config[col].tolist() for col in ["xpath", "auth_file", "auth_section", "auth_col", "separator"]
            )
        except Exception as e:
            tqdm.write(f"Failed to extract configuration columns for '{config_name}'. Error: {e}")
            return config_name, df

        # Count cores
        num_workers = os.cpu_count() - cores_spare or 1

        # Prepare arguments
        all_args = [
            (i, xpath, auth_file, auth_section, auth_col, separator)
            for i, (xpath, auth_file, auth_section, auth_col, separator)
            in enumerate(zip(xpaths, auth_files, auth_sections, auth_cols, separators))
        ]

        # Split into batches
        max_batches = num_workers + 1
        batch_size = max(1, math.ceil(len(all_args) / max_batches))

        batches = [
            all_args[i : i + batch_size]
            for i in range(0, len(all_args), batch_size)
        ]

        # Dispatch batches in parallel
        with ProcessPoolExecutor(max_workers=num_workers) as executor:
            futures = [
                executor.submit(process_collection_batch, batch, xml_data, lookup_df_list, separator_map)
                for batch in batches
            ]
            for future in tqdm(as_completed(futures), total=len(futures),
                            desc=f"File '{config_name}'", position=bar_pos):
                batch_result = future.result()
                for i, col_data in batch_result.items():
                    df.iloc[:, i] = col_data

        # Defragment the DataFrame by concatenation
        df = pd.concat([df], ignore_index=True)

        # Sort collection data
        df = sort_collection_df(df)

    else:
        raise ValueError(f"Unsupported file_type: {file_type}")

    # Save the processed DataFrame to CSV and JSON files
    save_as(df, csv_output_dir, config_name, format="csv")
    save_as(df, json_output_dir, config_name, format="json")

    # Return the outputs
    return config_name, df

# Helper function to process batches of collection columns
def process_authority_batch(batch, xml_data):
    """
    Processes a batch of authority columns by extracting data using XPath.
    Args:
        batch (list): List of tuples containing index, XPath, and authority file name.
        xml_data (dict): Dictionary of authority XML files.
    Returns:
        dict: Dictionary of processed results.
    """
    out = {}
    for i, xpath, auth_file in batch:
        _, col_data = process_authority_column(
            i,
            xpath,
            auth_file,
            xml_data
        )
        out[i] = col_data
    return out

# Helper function to process authority columns
def process_authority_column(i, xpath, auth_file, authority):
    """
    Processes a single column for the authority DataFrame.
    Args:
        i (int): The index of the column.
        xpath (str): The XPath expression to extract data.
        auth_file (str): The authority file name.
        authority (dict): Dictionary of authority XML files.
    Returns:
        tuple: The index and the extracted results.
    Outputs:
        csv: The extracted data is saved to a CSV file.
        json: The extracted data is saved to a JSON file.
    """
    auth_xml = authority.get(auth_file)
    results = extract_with_xpath(auth_xml, xpath)

    # Flatten the results and ensure no nested lists remain
    results = [item for sublist in results for item in (sublist if isinstance(sublist, list) else [sublist])]
    results = [item for item in results if not isinstance(item, list)]
    return i, results

# Helper function to process batches of collection columns
def process_collection_batch(batch, xml_data, lookup_df_list, separator_map):
    """
    Processes a batch of collection columns by extracting data using XPath and looking up values in an authority file.
    Args:
        batch (list): List of tuples containing index, XPath, authority file name, section, column, and separator.
        xml_data (dict): Dictionary of collection XML files.
        lookup_df_list (dict): Dictionary of DataFrames for authority files.
        separator_map (dict): Dictionary of separators for authority lookups.
    Returns:
        dict: Dictionary of processed results.
    """
    out = {}
    for i, xpath, auth_file, auth_section, auth_col, separator in batch:
        _, col_data = process_collection_column(
            i,
            xpath,
            auth_file,
            xml_data,
            lookup_df_list,
            auth_section,
            auth_col,
            separator,
            separator_map
        )
        out[i] = col_data
    return out

# Helper function to process collection columns
def process_collection_column(i, xpath, auth_file, catalogue, auth_df_list, auth_section, auth_col, separator, separator_map):
    """
    Processes a collection column by extracting data using XPath and looking up values in an authority file.
    Args:
        i (int): The index of the column.
        xpath (str): The XPath expression to extract data.
        auth_file (str): The authority file name.
        catalogue (dict): Dictionary of collection XML files.
        auth_df_list (dict): Dictionary of DataFrames for authority files.
        auth_section (str): The section name in the authority file.
        auth_col (str): The column name in the authority file.
        separator (str): The separator for joining values.
        separator_map (dict): Dictionary of separators for authority lookups.
    Returns:
        tuple: The index and the processed results.
    """
    # Set up list
    results = []

    # If auth_file is not in auth_df_list keys, extract the data and append directly
    if auth_file is None or auth_file.lower().strip() not in auth_df_list.keys():
        for filename, xml in catalogue.items():
            results.append(extract_with_xpath(xml, xpath))

    # Else extract the data and lookup in the authority DataFrame
    else:
        # Set up auth lookup DataFrame
        auth_df = auth_df_list.get(auth_file.lower().strip())

        # Set the separator
        s = get_separator(separator, separator_map)

        # Set the column name
        col_name = auth_section + ": " + auth_col

        # Lookup the value in the authority file
        for filename, xml in catalogue.items():
            # Build the lookup_data list using the helper function for each data_item
            lookup_data = [
                process_lookup_item(data_item, auth_df, col_name, s)
                for data_item in extract_with_xpath(xml, xpath)
            ]

            results.append(lookup_data)

    # Flatten the results and ensure no nested lists remain
    results = [item for sublist in results for item in (sublist if isinstance(sublist, list) else [sublist])]
    results = [item for item in results if not isinstance(item, list)]
    return i, results

# Helper function to apply XPath 2.0 queries to an XML element in the TEI namespace
def extract_with_xpath(xml_element, xpath_expr):
    """
    Extracts data from an XML element using XPath 2.0 queries.
    Args:
        xml_element (Element): The XML element to search.
        xpath_expr (str): The XPath expression to evaluate.
    Returns:
        list: The extracted data.
    """
    try:
        result = elementpath.select(
            xml_element, 
            xpath_expr, 
            namespaces={'tei': 'http://www.tei-c.org/ns/1.0'}
        )
        # Convert non-list results (including booleans) to a list.
        if not isinstance(result, list):
            result = [result]
    except Exception as e:
        tqdm.write(f"XPath extraction failed. Offending XPath: {xpath_expr}. Error: {e}")
        result = ""
    return result

# Helper function to determine the separator for authority lookups
def get_separator(separator, separator_map):
    if separator_map is None:
        s = "; "
        tqdm.write(f"No separator map found. Using '{s}' instead.")
    elif str(separator).lower().strip() in separator_map:
        s = separator_map.get(str(separator).lower().strip())
    elif "default" in separator_map:
        s = separator_map.get("default")
        tqdm.write(f"Encountered unexpected separator '{separator}'. Using default '{s}' instead.")
    else:
        s = "; "
        tqdm.write(f"Encountered unexpected separator '{separator}' and no default found. Using '{s}' instead.")
    return s

# Helper function to process data found through the authority lookup
def process_lookup_item(data_item, auth_df, col_name, separator):
    """
    Processes a single data item by looking it up in the authority DataFrame and returning the corresponding value.
    Args:
        data_item (str): The data item to process.
        auth_df (DataFrame): The authority DataFrame.
        col_name (str): The column name in the authority DataFrame.
        separator (str): The separator for joining values.
    Returns:
        str: The processed value, joined by the separator.
    """
    pieces = []
    # Split the data_item on spaces
    for identifier in data_item.split(" "):
        # Filter the DataFrame rows where the first column equals the identifier
        filtered = auth_df[auth_df.iloc[:, 0] == identifier]
        if not filtered.empty:
            # Get the value from the specified column
            value = filtered[col_name].iloc[0]
            # If the value is a boolean, convert its string form to lowercase
            piece = str(value).lower() if isinstance(value, bool) else str(value)
        else:
            piece = ""
        # Add only non-empty strings
        if piece:
            pieces.append(piece)
    
    # Deduplicate preserving the order by leveraging dict.fromkeys
    deduped = list(dict.fromkeys(pieces))
    # If all strings were empty, deduped will be empty; return a single empty string
    return separator.join(deduped) if deduped else ""

# Helper function to sort authority data
def sort_authority_df(df):
    """
    Sorts the authority DataFrame based on the first column.
    Args:
        df (DataFrame): The DataFrame to sort.
    Returns:
        DataFrame: The sorted DataFrame.
    """
    try:
        # If the first column contains numbers after "_", sort by that number
        if df.iloc[:, 0].str.contains(r'_\d+', na=False).any():
            df['temp'] = df.iloc[:, 0].str.extract(r'_(\d+)', expand=False).astype(float)
            df = df.sort_values(by='temp', ascending=True, na_position='last').reset_index(drop=True)
            df.drop(columns='temp', inplace=True)
        # Otherwise, sort by the first column directly
        else:
            df = df.sort_values(by=df.columns[0], ascending=True, na_position='last').reset_index(drop=True)
        return df
    except Exception as e:
        tqdm.write(f"Sorting DataFrame failed. Error: {e}")
        return df

# Helper function to sort collection data
def sort_collection_df(df):
    """
    Sorts the collection DataFrame based on the first column or 'file URL'.
    Args:
        df (DataFrame): The DataFrame to sort.
    Returns:
        DataFrame: The sorted DataFrame.
    """
    try:
        # If 'metadata: file URL' exists, sort by it first, then by the first column
        if 'metadata: file URL' in df.columns:
            # Extract and sort by numeric part in the 'file URL'.
            df['metadata: file URL temp'] = df['metadata: file URL'].str.extract(r'manuscript_(\d+)')[0].astype(float)
            first_col = df.columns[0]
            sort_by = ['metadata: file URL temp'] if first_col == 'metadata: file URL' else ['metadata: file URL temp', first_col]
            df.sort_values(by=sort_by, ascending=True, na_position='last', inplace=True)
            df.drop(columns=['metadata: file URL temp'], inplace=True)
        # Otherwise, if 'metadata: collection' exists, sort by it first, then by the first column
        elif 'metadata: collection' in df.columns:
            # If a 'metadata: collection' column exists, sort by it, then natural sort on the first column.
            first_col = df.columns[0]
            df.sort_values(
                by=['metadata: collection', first_col],
                key=lambda col: col.map(natural_keys),
                ascending=True,
                na_position='last',
                inplace=True
            )
        # Otherwise, sort by the first column directly
        else:
            # Otherwise, use a natural sort on the first column.
            first_col = df.columns[0]
            df.sort_values(
                by=first_col,
                key=lambda col: col.map(natural_keys),
                ascending=True,
                na_position='last',
                inplace=True
            )
        return df
    except Exception as e:
        tqdm.write(f"Sorting DataFrame failed. Error: {e}")
        return df

# Helper function for natural sorting of strings
def natural_keys(text):
    """
    Convert a string into a list of integers and strings for natural sorting.
    Args:
        text (str): The string to convert.
    Returns:
        list: A list of integers and strings.
    """
    list = [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text)]
    return list

# Helper function to save DataFrame as either csv or json file
def save_as(df, output_dir, config_name, format):
    """
    Saves a DataFrame to a file in the specified format.
    Args:
        df (DataFrame): The DataFrame to save.
        output_dir (str): Directory to save the file.
        config_name (str): Name of the configuration file.
        format (str): File format to save as. Must be 'csv' or 'json'.
    """
    os.makedirs(output_dir, exist_ok=True)
    output_filename = f"{config_name}.{format}"
    output_file = os.path.join(output_dir, output_filename)
    try:
        if format == "csv":
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
            tqdm.write(f"Saved '{config_name}' to '{output_file}'")
        elif format == "json":
            df.to_json(output_file, orient='records', lines=True, force_ascii=False)
            tqdm.write(f"Saved '{config_name}' to '{output_file}'")
        else:
            tqdm.write(f"Invalid format '{format}'. Supported formats are 'csv' and 'json'.")
    except Exception as e:
        tqdm.write(f"Saving data to '{output_filename}' failed. Error: {e}")

# Function to save DataFrame list as an xlsx file with individual tables as tabs
def save_as_xlsx(df_list, config_list, output_dir, output_filename):
    """
    Saves a list of DataFrames to an Excel file with each DataFrame in a separate sheet.
    Args:
        df_list (dict): Dictionary of DataFrames to save.
        config_list (dict): Dictionary of configuration DataFrames for headings and sections.
        output_dir (str): Directory to save the Excel file.
        output_filename (str): Name of the output Excel file (without extension).
    """
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{output_filename}.xlsx")
    sections_list = [config_list[config_name]['section'].to_numpy() for config_name in config_list.keys()]
    headings_list = [config_list[config_name]['heading'].to_numpy() for config_name in config_list.keys()]
    comments_list = [config_list[config_name]['comment'].to_numpy() for config_name in config_list.keys()]
    tqdm.write(f"Saving '{output_filename}'...")
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Zip over the sheets' data, section titles, and comments
            for (name, df), sections, headings, comments in zip(df_list.items(), sections_list, headings_list, comments_list):
                # Convert numpy_bool to bool
                for col in df.select_dtypes(include="bool"):
                    df[col] = df[col].astype(bool)
                
                # Write the DataFrame starting from row 2
                df.to_excel(writer, sheet_name=name, index=False, startrow=1)

                # Set the values in row 2 to headings
                for col_idx, value in enumerate(headings, start=1):
                    writer.sheets[name].cell(row=2, column=col_idx, value=value)

                # Access the workbook and worksheet
                worksheet = writer.sheets[name]

                # Write the section titles into the first row
                for col_idx, value in enumerate(sections, start=1):
                    worksheet.cell(row=1, column=col_idx, value=value)

                # Force text type if value starts with '='
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            cell.value = "'" + cell.value

                # Set each column to the width of the content in the second row, with a minimum value
                for col_idx, cell in enumerate(worksheet[2], start=1):
                    column_letter = get_column_letter(col_idx)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length >= 10:
                            worksheet.column_dimensions[column_letter].width = cell_length
                        else:
                            worksheet.column_dimensions[column_letter].width = 10
                    else:
                        worksheet.column_dimensions[column_letter].width = 10

                # Add comments to each cell of the second row using the relevant value from comments
                for col_idx, comment_text in enumerate(comments, start=1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    comment = Comment(comment_text, "Generated")
                    # Set comment height, assuming 15 characters per line and 15pt per line
                    num_lines = (len(str(comment_text)) // 15)
                    comment.height = 30 + 15 * num_lines
                    # Set comment width to a default value
                    comment.width = 200
                    # Set the comment to the cell
                    cell.comment = comment

                # Set up a filter for each column, with row 2 given as the header value
                last_row = worksheet.max_row
                last_col_letter = get_column_letter(len(sections))
                worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

                # Merge and center identical consecutive section values in the first row
                merge_and_center_cells(worksheet, sections)

                # Freeze the first two rows
                worksheet.freeze_panes = worksheet['A3']

        tqdm.write(f"Saved data to '{output_filename}'")

    except Exception as e:
        tqdm.write(f"Saving data to '{output_filename}' failed. Error: {e}")

# Helper function to merge and center identical consecutive section values in the first row of an xlsx file
def merge_and_center_cells(worksheet, sections):
    """
    Merges and centers identical consecutive section values in the first row of the worksheet.

    Args:
        worksheet: The worksheet object where merging is applied.
        sections: A list of section titles corresponding to the columns.
    """
    start_col = 1
    for col_idx in range(1, len(sections) + 1):
        if col_idx == len(sections) or sections[col_idx] != sections[start_col - 1]:
            if col_idx - start_col >= 1:
                worksheet.merge_cells(
                    start_row=1, start_column=start_col,
                    end_row=1, end_column=col_idx
                )
                merged_cell = worksheet.cell(row=1, column=start_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                worksheet.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
            start_col = col_idx + 1
